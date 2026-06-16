"""
Reference verification harness
==============================
Diffs the engine output against the hand-built reference schedule in
`For Python_final.xlsx` (sheet B+S+B+S), column-by-column, row-by-row.

The reference is the source of truth: we read its market-price column (L)
and feed those exact quotes to the engine, so the only thing under test is
the engine's own math (amortization, carrying value, accrued, MTM, OCI, NAV).

Run:  set PYTHONIOENCODING=utf-8 && python verify_ref.py
"""
import sys, os
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from engine.calculator import build_cashflow_table

REF_FILE = 'For Python_final.xlsx'
SHEET    = 'B+S+B+S'

# Reference column letters -> (engine row field, tolerance)
COLS = [
    ('E', 'bond_discount',   1.0),   # remaining discount
    ('F', 'cum_amort_disc',  1.0),   # cumulative amortized discount
    ('G', 'bond_premium',    1.0),
    ('H', 'cum_amort_prem',  1.0),
    ('I', 'carrying_value',  1.0),   # ending carrying value
    ('J', 'accrued_int',     1.0),   # accrued interest
    ('K', 'nominal_balance', 1.0),   # nominal
    ('L', 'price',           1e-4),  # market price
    ('M', 'mtm',             1.0),
    ('N', 'oci_gl',          1.0),
    ('O', 'nav',             1.0),
    ('P', 'check',           0.01),
    ('V', 'default_prob',    1e-6),  # DRSK default probability (input)
    ('W', 'lgd',             1e-6),  # LGD (input)
    ('X', 'expected_loss',   1.0),   # nominal × PD × LGD
    ('Y', 'el_change',       1.0),   # period-over-period change in EL
]
COL_IDX = {c: i for i, c in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ', 1)}


def read_reference():
    """Return {date: {colletter: value}} for the dated schedule rows, plus the
    market-price series {date: price} taken straight from column L."""
    wb = openpyxl.load_workbook(REF_FILE, data_only=True)
    ws = wb[SHEET]
    ref, prices = {}, {}
    for r in range(33, 77):
        a = ws.cell(row=r, column=1).value
        if not isinstance(a, datetime):
            continue
        d = a.date()
        ref[d] = {c: ws.cell(row=r, column=COL_IDX[c]).value for c, _f, _t in COLS}
        L = ws.cell(row=r, column=COL_IDX['L']).value
        if L is not None:
            prices[d] = float(L)
    return ref, prices


def build_params(prices):
    return dict(
        isin='US91', issuer='', par_value=137_000_000.0, clean_price=99.86328125,
        coupon_rate=4.0, interest_frequency=2,
        settle_date=date(2026, 5, 14), last_interest_date=date(2026, 1, 31),
        next_interest_date=date(2026, 7, 31), maturity_date=date(2029, 7, 31),
        default_probability=0.000682, lgd=0.6,   # DRSK / LGD inputs (reference V/W)
        market_prices=prices,
        transactions=[
            {'date': date(2026, 11, 30), 'type': 'BUY',          'nominal': 5_900_000.0, 'clean_price': 99.640625, 'accrued_interest': 0},
            {'date': date(2027, 11, 30), 'type': 'SELL_PARTIAL', 'nominal': 5_900_000.0, 'clean_price': 97.132788, 'accrued_interest': 0},
            {'date': date(2028, 4, 30),  'type': 'BUY',          'nominal': 5_900_000.0, 'clean_price': 99.62,     'accrued_interest': 0},
            {'date': date(2028, 9, 30),  'type': 'SELL_PARTIAL', 'nominal': 6_000_000.0, 'clean_price': 97.132788, 'accrued_interest': 0},
        ])


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def main():
    ref, prices = read_reference()
    rows = build_params(prices)
    eng = {r['date']: r for r in build_cashflow_table(rows) if r.get('date')}

    common = sorted(set(ref) & set(eng))
    print(f'reference dated rows: {len(ref)} | engine dated rows: {len(eng)} | common: {len(common)}')
    print()
    print(f'{"col":3} {"field":18} {"match":>7} {"total":>7}   first mismatches (date: engine vs ref)')
    total_bad = 0
    for c, fld, tol in COLS:
        m = t = 0
        bad = []
        for d in common:
            rv = num(ref[d].get(c))
            if rv is None:
                continue
            ev = num(eng[d].get(fld))
            t += 1
            if ev is not None and abs(ev - rv) <= tol:
                m += 1
            elif len(bad) < 3:
                bad.append(f'{d}: {ev if ev is not None else "None":>14} vs {rv:,.2f}'
                           if ev is None else f'{d}: {ev:,.2f} vs {rv:,.2f}')
        total_bad += (t - m)
        flag = 'OK ' if m == t else '!! '
        print(f'{flag}{c:2} {fld:18} {m:>7} {t:>7}   {" | ".join(bad)}')
    print()
    print('ALL COLUMNS MATCH' if total_bad == 0 else f'{total_bad} mismatched cells remain')


if __name__ == '__main__':
    main()
