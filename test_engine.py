"""Sanity tests for the bond engine + excel round-trip. Run: python test_engine.py"""
import io
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from datetime import date

from engine.calculator import (
    build_cashflow_table, get_summary, calc_ytm, validate_params,
    derive_accrued_interest, parse_date, safe_float
)
from engine.excel_io import export_to_excel, read_from_excel

FAILS = []
def check(name, cond, detail=''):
    print(f"{'PASS' if cond else 'FAIL'}  {name}  {detail}")
    if not cond:
        FAILS.append(name)

# ── Base bond: 1,000,000 par, 8% semi-annual, bought at 95 ──────────────────
base = dict(
    isin='TEST01', issuer='Test Issuer',
    par_value=1_000_000.0, clean_price=95.0, coupon_rate=8.0,
    interest_frequency=2,
    settle_date=date(2024,1,15), last_interest_date=date(2023,7,15),
    next_interest_date=date(2024,7,15), maturity_date=date(2029,1,15),
)

# 1. Simple hold-to-maturity
rows = build_cashflow_table(dict(base))
s = get_summary(dict(base), rows)
cpp = 1_000_000 * 0.08 / 2  # 40,000 per period

coupon_rows = [r for r in rows if r.get('is_coupon') and not r.get('is_maturity')]
check("coupon rows exist", len(coupon_rows) == 9, f"({len(coupon_rows)} found, expect 9)")
check("coupon cashflow = 40,000", all(abs(r['cashflow'] - cpp) < 0.01 for r in coupon_rows))
check("accrued = full coupon on coupon date (matches cashflow)",
      all(abs(r['accrued_int'] - cpp) < 0.01 for r in coupon_rows))

mat = [r for r in rows if r.get('is_maturity')][0]
check("maturity cashflow = principal + final coupon",
      abs(mat['cashflow'] - (1_000_000 + cpp)) < 0.01, f"(got {mat['cashflow']:,.2f})")
check("maturity carrying value accreted to par",
      abs(mat['carrying_value'] - 1_000_000) < 0.01, f"(got {mat['carrying_value']:,.2f})")
check("maturity MTM zeroed", mat['mtm'] == 0.0)

# Accrued at purchase: 184 days into a 182-day period? 15Jul23→15Jan24 = 184d, settle 15Jan24 → full period
acc0 = derive_accrued_interest(1_000_000, 8.0, 2, date(2024,1,15), date(2023,7,15), date(2024,1,15))
# settle == coupon date scenario isn't ours; engine derives vs 15Jul24 next:
acc = derive_accrued_interest(1_000_000, 8.0, 2, date(2024,4,15), date(2024,1,15), date(2024,7,15))
check("accrued mid-period ≈ half coupon", abs(acc - cpp * 91/182) < 1.0, f"(got {acc:,.2f})")

# Net cashflow identity: -invested + coupons + redemption = net
expect_net = s['total_coupon_income'] + s['redemption'] + s['total_sell_proceeds'] - s['total_invested']
check("net cashflow identity", abs(s['net_cashflow'] - expect_net) < 0.01,
      f"(net {s['net_cashflow']:,.2f} vs {expect_net:,.2f})")
check("status = Matured", s['status'] == 'Matured')

# 2. YTM: discount bond YTM > coupon; par bond YTM == coupon
ytm = s['ytm']
check("YTM computed", ytm is not None, f"(YTM={ytm})")
check("discount bond: YTM > coupon", ytm is not None and ytm > 8.0, f"(YTM={ytm})")
par_params = dict(base, clean_price=100.0, settle_date=date(2024,7,15),
                  last_interest_date=date(2024,1,15), next_interest_date=date(2025,1,15),
                  accrued_interest=0.0)
ytm_par = get_summary(par_params, build_cashflow_table(par_params))['ytm']
check("par bond on coupon date: YTM ≈ coupon", ytm_par is not None and abs(ytm_par - 8.0) < 0.01,
      f"(YTM={ytm_par})")

# 3. Multi-transaction: buy more, partial sell, auto accrued
multi = dict(base, transactions=[
    {'date': date(2024,9,15), 'type': 'BUY', 'nominal': 500_000.0,
     'clean_price': 96.0, 'accrued_interest': 0, 'note': 'tap'},
    {'date': date(2025,3,15), 'type': 'SELL_PARTIAL', 'nominal': 600_000.0,
     'clean_price': 97.0, 'accrued_interest': 0, 'note': 'trim'},
])
rows_m = build_cashflow_table(multi)
s_m = get_summary(multi, rows_m)
buy2 = [r for r in rows_m if r.get('is_buy') and not r.get('is_header')][0]
# auto-accrued for buy on 15-Sep-24: last cpn 15-Jul-24, next 15-Jan-25 (184d), 62 days elapsed
exp_acc = 500_000*0.08/2 * 62/184
check("BUY auto-accrued derived", abs(buy2['accrued_int'] - exp_acc) < 0.5,
      f"(got {buy2['accrued_int']:,.2f}, expect {exp_acc:,.2f})")
sell = [r for r in rows_m if r.get('is_sell')][0]
exp_sell_acc = 600_000*0.08/2 * 59/181  # 15-Jan-25→15-Mar-25 = 59d of 181d period
check("SELL auto-accrued derived", abs(sell['accrued_int'] - exp_sell_acc) < 0.5,
      f"(got {sell['accrued_int']:,.2f}, expect {exp_sell_acc:,.2f})")
# Transactions are recognised on their own date: the sell row shows the reduced
# position (1.5m − 600k = 900k) directly, with proceeds and realized P&L.
check("balance after partial sell = 900,000", sell['nominal_balance'] == 900_000,
      f"(got {sell['nominal_balance']:,.0f})")
check("realized P&L present", sell['realized_pl'] is not None and s_m['total_realized_pl'] == sell['realized_pl'])
# coupons after the buy reflect 1.5m nominal
cpn_after_buy = [r for r in rows_m if r.get('is_coupon') and not r.get('is_maturity')
                 and r['date'] == date(2025,1,15)][0]
check("coupon after buy = 60,000", abs(cpn_after_buy['cashflow'] - 60_000) < 0.01,
      f"(got {cpn_after_buy['cashflow']:,.2f})")
check("maturity pays remaining 900,000 + coupon",
      abs([r for r in rows_m if r.get('is_maturity')][0]['cashflow'] - (900_000 + 36_000)) < 0.01)

# 4. SELL_FULL with zero nominal (previously dropped) + tx after maturity ignored
full = dict(base, transactions=[
    {'date': date(2025,6,1), 'type': 'SELL_FULL', 'nominal': 0,
     'clean_price': 98.0, 'accrued_interest': 0, 'note': ''},
    {'date': date(2030,1,1), 'type': 'BUY', 'nominal': 100_000,
     'clean_price': 99.0, 'accrued_interest': 0, 'note': 'bad date'},
])
rows_f = build_cashflow_table(full)
s_f = get_summary(full, rows_f)
sells_f = [r for r in rows_f if r.get('is_sell')]
check("SELL_FULL w/o nominal sells everything", len(sells_f) == 1 and sells_f[0]['nominal_change'] == -1_000_000)
check("no rows after full sell-out", rows_f[-1]['date'] == date(2025,6,1) or not any(
      r.get('date') and r['date'] > date(2025,6,1) for r in rows_f))
check("tx after maturity ignored", not any(r.get('date') and r['date'] > base['maturity_date'] for r in rows_f))
check("status = Closed", s_f['status'] == 'Closed', f"(got {s_f['status']})")
warns = validate_params(full)
check("validation flags post-maturity tx", any('after maturity' in w for w in warns), f"({warns})")

# 5. Oversell warning + clamping
over = dict(base, transactions=[
    {'date': date(2025,6,1), 'type': 'SELL_PARTIAL', 'nominal': 2_000_000.0,
     'clean_price': 98.0, 'accrued_interest': 0, 'note': ''}])
warns_o = validate_params(over)
check("validation flags oversell", any('exceeds holding' in w for w in warns_o))
rows_o = build_cashflow_table(over)
check("oversell clamped to holding", [r for r in rows_o if r.get('is_sell')][0]['nominal_change'] == -1_000_000)

# 6. OCI consistency: discount bond oci == mtm (no double count)
snap = [r for r in rows if r.get('date') == date(2026,1,15)][0]
check("OCI G/L == MTM (consistent disc/prem)", snap['oci_gl'] == snap['mtm'])

# 7. parse helpers
check("safe_float parses '95.5'", safe_float('95.5') == 95.5)
check("safe_float parses '1,000,000'", safe_float('1,000,000') == 1_000_000)
check("parse_date '15.01.2024'", parse_date('15.01.2024') == date(2024,1,15))

# 8. Excel export + round-trip re-import
xls = export_to_excel(multi, rows_m)
check("excel export produces bytes", isinstance(xls, bytes) and len(xls) > 5000, f"({len(xls)} bytes)")
re_params, re_errors = read_from_excel(io.BytesIO(xls))
check("re-import has no errors", not re_errors, f"({re_errors})")
check("round-trip par value", re_params.get('par_value') == 1_000_000)
check("round-trip clean price", re_params.get('clean_price') == 95.0)
check("round-trip dates", re_params.get('settle_date') == date(2024,1,15)
      and re_params.get('maturity_date') == date(2029,1,15)
      and re_params.get('last_interest_date') == date(2023,7,15)
      and re_params.get('next_interest_date') == date(2024,7,15))
check("round-trip transactions", len(re_params.get('transactions', [])) == 2,
      f"({len(re_params.get('transactions', []))} txs)")
if len(re_params.get('transactions', [])) == 2:
    t1, t2 = re_params['transactions']
    check("round-trip tx types", t1['type'] == 'BUY' and t2['type'] == 'SELL_PARTIAL',
          f"({t1['type']}, {t2['type']})")
    check("round-trip tx nominals", t1['nominal'] == 500_000 and t2['nominal'] == 600_000)
# recalculate from re-imported params and compare realized P&L
rows_rt = build_cashflow_table(re_params)
s_rt = get_summary(re_params, rows_rt)
check("round-trip recalculation matches realized P&L",
      abs(s_rt['total_realized_pl'] - s_m['total_realized_pl']) < 0.01,
      f"({s_rt['total_realized_pl']:,.2f} vs {s_m['total_realized_pl']:,.2f})")

# 9. Excel nominal column actually written (the old bug)
import openpyxl
wb = openpyxl.load_workbook(io.BytesIO(xls))
ws = wb['Bond Report']
hdr = [ws.cell(row=4, column=i).value for i in range(1, 27)]
check("excel has P&L (Sell) column", 'P&L (Sell)' in hdr)
nom_col = hdr.index('Nominal Δ') + 1
check("excel Nominal Δ on buy row = 1,000,000", ws.cell(row=5, column=nom_col).value == 1_000_000)
check("excel Summary sheet present", 'Summary' in wb.sheetnames)

# 10. SELL_FULL parsed correctly from excel text
full_x = dict(base, transactions=[{'date': date(2025,6,1), 'type': 'SELL_FULL',
    'nominal': 1_000_000.0, 'clean_price': 98.0, 'accrued_interest': 0, 'note': ''}])
xls2 = export_to_excel(full_x, build_cashflow_table(full_x))
re2, _ = read_from_excel(io.BytesIO(xls2))
check("SELL_FULL survives round-trip", re2['transactions'][0]['type'] == 'SELL_FULL',
      f"(got {re2['transactions'][0]['type']})")

# ── 11. Multi-buy: the buy row (recognised on its date) shows the FULL combined
#        position, and every later row carries it forward. ────────────────────
# lot1: 1,000,000 @95 from 15-Jan-24 (disc -50,000, 1827 days to maturity)
# buy2: 500,000 @96 on 15-Sep-24 (244 days elapsed on lot1, disc -20,000)
lot1_cv_at_buy2 = 1_000_000 - 50_000 * (1 - 244 / 1827)
exp_cv = lot1_cv_at_buy2 + 500_000 * 0.96
check("buy row shows combined carrying value",
      abs(buy2['carrying_value'] - exp_cv) < 0.01,
      f"(got {buy2['carrying_value']:,.2f}, expect {exp_cv:,.2f})")
exp_disc = -50_000 * (1583 / 1827) + (-20_000)   # lot1 remaining + lot2 full
check("buy row shows combined remaining discount",
      abs(buy2['bond_discount'] - exp_disc) < 0.01,
      f"(got {buy2['bond_discount']:,.2f}, expect {exp_disc:,.2f})")
check("buy row shows combined nominal = 1,500,000", buy2['nominal_balance'] == 1_500_000,
      f"(got {buy2['nominal_balance']:,.0f})")
check("buy row NAV/MTM populated", buy2['nav'] is not None and buy2['mtm'] is not None)
check("sell row shows reduced position (900,000, positive carrying)",
      sell['nominal_balance'] == 900_000 and sell['carrying_value'] > 0,
      f"(nom={sell['nominal_balance']}, cv={sell['carrying_value']})")
# carrying continuity: buy-row carrying ≈ next period-row carrying (same date logic)
next_period = next(r for r in rows_m if r.get('date') and r['date'] > buy2['date']
                   and not r.get('is_buy') and not r.get('is_sell'))
check("no jump after buy (continuity)",
      abs(next_period['carrying_value'] - buy2['carrying_value']) < 1500,
      f"(buy {buy2['carrying_value']:,.2f} → next {next_period['carrying_value']:,.2f})")

# ── 12. Check columns tie out to ~0 everywhere ────────────────────────────────
bad_check = [r for r in rows_m if r.get('check') not in (None, 0) and abs(r['check']) > 0.02]
bad_check2 = [r for r in rows_m if r.get('check2') not in (None, 0) and abs(r['check2']) > 0.02]
check("Check column = 0 on all rows", not bad_check, f"({len(bad_check)} bad)")
check("Check 2 column = 0 on all rows", not bad_check2, f"({len(bad_check2)} bad)")

# ── 13. Realized interest income & Total P&L running totals ──────────────────
s_m2 = get_summary(multi, build_cashflow_table(multi))
acc0_paid = derive_accrued_interest(1_000_000, 8.0, 2, date(2024,1,15), date(2023,7,15), date(2024,7,15))
exp_rii = (s_m2['total_coupon_income']            # coupons + final coupon received
           - acc0_paid                            # accrued paid at initial buy
           - buy2['accrued_int']                  # accrued paid at buy #2
           + sell['accrued_int'])                 # accrued received at sell
check("realized interest income = coupons - accrued paid + accrued received",
      abs(s_m2['realized_interest_income'] - exp_rii) < 0.5,
      f"(got {s_m2['realized_interest_income']:,.2f}, expect {exp_rii:,.2f})")
check("total P&L = sells P&L + interest income",
      abs(s_m2['total_pl'] - (s_m2['total_realized_pl'] + s_m2['realized_interest_income'])) < 0.01)
last_row = [r for r in rows_m if r.get('total_pl') is not None][-1]
check("total P&L is a running column", last_row['total_pl'] == s_m2['total_pl'])

# ── 14. Credit risk: Expected Loss = NOMINAL × PD × LGD (reference X col), Change=ΔEL ─
credit = dict(base, default_probability=2.0, lgd=60.0)   # % inputs normalize
rows_c = build_cashflow_table(credit)
mid = [r for r in rows_c if r.get('date') == date(2026,1,15)][0]
exp_el = round(1_000_000 * 0.02 * 0.60, 2)   # nominal exposure 1m
check("expected loss = nominal × PD × LGD", abs(mid['expected_loss'] - exp_el) < 0.01,
      f"(got {mid['expected_loss']:,.2f}, expect {exp_el:,.2f})")
# EL is flat between position changes; the first credit row shows the full
# initial EL and the maturity row drops it to zero (bond redeemed).
el_rows = [r for r in rows_c if r.get('el_change') is not None]
check("EL flat between first row and maturity (no txs)",
      all(r['el_change'] == 0 for r in el_rows[1:-1]),
      f"(non-flat: {[r.get('date') for r in el_rows[1:-1] if r['el_change'] != 0]})")
check("first credit row shows full initial EL", abs(el_rows[0]['el_change'] - exp_el) < 0.01,
      f"(got {el_rows[0]['el_change']:,.2f})")
check("EL drops to zero at maturity", el_rows[-1]['expected_loss'] == 0,
      f"(got {el_rows[-1]['expected_loss']})")
check("PD/LGD as fractions on every row",
      all(r['default_prob'] == 0.02 and r['lgd'] == 0.6 for r in rows_c))
rows_cf = build_cashflow_table(dict(base, default_probability=0.02, lgd=0.6))
check("fraction inputs pass through", rows_cf[0]['default_prob'] == 0.02 and rows_cf[0]['lgd'] == 0.6)
# PD auto-derived from price when not provided (credit triangle)
rows_auto = build_cashflow_table(dict(base))
exp_pd = (1 - 0.95) / (0.6 * 1827/365.25)
check("PD implied from price discount", abs(rows_auto[0]['default_prob'] - exp_pd) < 1e-6,
      f"(got {rows_auto[0]['default_prob']:.8f}, expect {exp_pd:.8f})")
check("implied PD: premium bond → 0",
      build_cashflow_table(dict(base, clean_price=102.0))[0]['default_prob'] == 0.0)
# EL drops when half the position is sold — shown on the sell row itself
# (recognised on its date): EL halves and the Change is negative.
rows_cs = build_cashflow_table(dict(credit, transactions=[
    {'date': date(2025,6,1), 'type': 'SELL_PARTIAL', 'nominal': 500_000.0,
     'clean_price': 98.0, 'accrued_interest': 0, 'note': ''}]))
sell_c = [r for r in rows_cs if r.get('is_sell')][0]
check("EL halves after selling half", abs(sell_c['expected_loss'] - exp_el/2) < 1.0,
      f"(got {sell_c['expected_loss']:,.2f})")
check("EL Change negative on sell", sell_c['el_change'] < 0,
      f"(got {sell_c['el_change']})")

# ── 15. Export carries new columns + PD/LGD round-trip ────────────────────────
xls_c = export_to_excel(credit, rows_c)
wb_c = openpyxl.load_workbook(io.BytesIO(xls_c))
hdr_c = [wb_c['Bond Report'].cell(row=4, column=i).value for i in range(1, 27)]
for col in ['Check','P&L (Sell)','Realized Interest Income','Total P&L','WAC',
            'Check 2','Default probability (DRSK)','LGD (1-CDS recovery)',
            'Expected Loss','Change']:
    check(f"excel column '{col}' present", col in hdr_c)
re_c, re_c_err = read_from_excel(io.BytesIO(xls_c))
re_rows = build_cashflow_table(re_c)
check("PD/LGD round-trip (normalized)",
      re_rows[0]['default_prob'] == 0.02 and re_rows[0]['lgd'] == 0.6,
      f"(PD={re_rows[0]['default_prob']}, LGD={re_rows[0]['lgd']})")
# recovery-rate label derives LGD in the engine
from openpyxl import Workbook
wb_r = Workbook(); ws_r = wb_r.active
for i, (l, v) in enumerate([('Par Value', 1_000_000), ('Clean Price', 95.0),
    ('Coupon Rate', 8.0), ('Settle Date', date(2024,1,15)),
    ('Last Interest Date', date(2023,7,15)), ('Next Interest Date', date(2024,7,15)),
    ('Maturity Date', date(2029,1,15)), ('Default Probability', 1.5),
    ('CDS Recovery Rate', 40.0)], 1):
    ws_r.cell(row=i, column=1, value=l); ws_r.cell(row=i, column=2, value=v)
buf_r = io.BytesIO(); wb_r.save(buf_r); buf_r.seek(0)
re_r, re_r_err = read_from_excel(buf_r)
rows_rr = build_cashflow_table(re_r)
check("LGD derived from recovery rate (1-0.4=0.6)", rows_rr[0]['lgd'] == 0.6,
      f"(got {rows_rr[0]['lgd']}, errors {re_r_err})")
check("explicit PD 1.5% normalized", rows_rr[0]['default_prob'] == 0.015)

# ── 16. fake.xlsx: multiple buys as repeated parameter blocks ─────────────────
fk, fk_err = read_from_excel('fake.xlsx')
check("fake.xlsx reads without errors", not fk_err, f"({fk_err})")
check("fake.xlsx par value = 137,000,000", fk.get('par_value') == 137_000_000)
check("fake.xlsx clean price = 99.86328125", fk.get('clean_price') == 99.86328125)
check("fake.xlsx dates", fk.get('settle_date') == date(2026,5,14)
      and fk.get('maturity_date') == date(2029,7,31)
      and fk.get('last_interest_date') == date(2026,1,31)
      and fk.get('next_interest_date') == date(2026,7,31))
check("fake.xlsx finds 2nd buy as transaction", len(fk.get('transactions', [])) == 1,
      f"({len(fk.get('transactions', []))} txs)")
if fk.get('transactions'):
    t2 = fk['transactions'][0]
    check("fake.xlsx buy #2 fields",
          t2['type'] == 'BUY' and t2['nominal'] == 5_900_000
          and t2['clean_price'] == 99.640625 and t2['date'] == date(2026,5,27)
          and abs(t2['accrued_interest'] - 75624.3094) < 0.01,
          f"({t2})")
rows_fk = build_cashflow_table(fk)
s_fk = get_summary(fk, rows_fk)
buy2_fk = [r for r in rows_fk if r.get('is_buy') and not r.get('is_header')]
check("fake.xlsx schedule has buy #2 row", len(buy2_fk) == 1)
# The buy #2 row (recognised on its date) shows the combined position directly.
check("fake.xlsx combined nominal = 142,900,000",
      buy2_fk[0]['nominal_balance'] == 142_900_000,
      f"(got {buy2_fk[0]['nominal_balance']:,.0f})")
check("fake.xlsx initial discount = -187,304.69",
      abs(rows_fk[0]['bond_discount'] - (-187_304.69)) < 0.01,
      f"(got {rows_fk[0]['bond_discount']:,.2f})")
mat_fk = [r for r in rows_fk if r.get('is_maturity')][0]
check("fake.xlsx maturity = 142.9m + 2,858,000 coupon",
      abs(mat_fk['cashflow'] - (142_900_000 + 2_858_000)) < 0.01,
      f"(got {mat_fk['cashflow']:,.2f})")
check("fake.xlsx coupon after buy2 = 2,858,000",
      abs(next(r['cashflow'] for r in rows_fk if r.get('is_coupon') and not r.get('is_maturity')
               and r['date'] == date(2026,7,31)) - 2_858_000) < 0.01)
check("fake.xlsx PD implied, LGD 0.6",
      rows_fk[0]['lgd'] == 0.6 and 0.0005 < rows_fk[0]['default_prob'] < 0.001,
      f"(PD={rows_fk[0]['default_prob']:.8f} vs sheet 0.000682)")
# EL is flat except: first credit row (full initial EL), the buy #2 row
# (position steps up), and maturity (redeemed → drops to zero).
el_rows_fk = [r for r in rows_fk if r.get('el_change') is not None]
check("fake.xlsx EL flat except the buy #2 row (interior)",
      all(r['el_change'] == 0 for r in el_rows_fk[1:-1] if r is not buy2_fk[0]),
      f"(unexpected: {[r.get('date') for r in el_rows_fk[1:-1] if r is not buy2_fk[0] and r['el_change'] != 0]})")
check("fake.xlsx EL steps up on buy #2", buy2_fk[0]['el_change'] > 0,
      f"(got {buy2_fk[0]['el_change']})")
check("fake.xlsx EL drops to zero at maturity", el_rows_fk[-1]['expected_loss'] == 0,
      f"(got {el_rows_fk[-1]['expected_loss']})")

# ── 17. Premium-bond sign check (self-contained): cumulative premium
#        amortization is NEGATIVE, and OCI = MTM − cum_amort_disc − cum_amort_prem
#        nets to MTM + premium_amortized. Discount bonds never exposed this. ─────
prem_bond = dict(isin='PREM1', par_value=1_000_000.0, clean_price=105.0, coupon_rate=8.0,
                 interest_frequency=2, settle_date=date(2024,1,15), last_interest_date=date(2023,7,15),
                 next_interest_date=date(2024,7,15), maturity_date=date(2029,1,15),
                 market_prices=[{'date': date(2024,1,15), 'price': 105.0}])  # flat at cost → MTM≈0
pr = [r for r in build_cashflow_table(prem_bond) if r.get('date') == date(2026,1,15)][0]
check("premium bond: bond_premium remaining > 0", pr['bond_premium'] > 0, f"(got {pr['bond_premium']:.2f})")
check("premium bond: cum_amort_prem is NEGATIVE", pr['cum_amort_prem'] < 0, f"(got {pr['cum_amort_prem']:.2f})")
# With market = cost, MTM = 0, so OCI must equal the premium amortized so far (positive)
prem_amortized = -pr['cum_amort_prem']
check("premium bond: OCI = MTM + premium amortized (sign correct)",
      abs(pr['oci_gl'] - (pr['mtm'] + prem_amortized)) < 0.01,
      f"(oci {pr['oci_gl']:.2f} vs mtm {pr['mtm']:.2f} + amort {prem_amortized:.2f})")
check("premium bond: OCI not negated (would be wrong sign)", pr['oci_gl'] > 0,
      f"(got {pr['oci_gl']:.2f})")

# ── 18. Calculator-page parity (clean automated reference 'exel caalculator',
#        a premium bond): the engine must reproduce EVERY valuation column exactly.
#        (The older B+S+B+S sheet is not asserted here: it mixes deferred and
#        immediate transaction recognition, whereas the engine recognises every
#        transaction on its own date — see test 11.)
import os as _os
if _os.path.exists('~$For Python_final.xlsx'):
    print("SKIP  file parity — For Python_final.xlsx is open in Excel (close it for a clean check)")
elif _os.path.exists('For Python_final.xlsx'):
    import verify_calc as vc
    _cref, _cpr = vc.read_reference()
    _ceng = {r['date']: r for r in build_cashflow_table(vc.build_params(_cpr)) if r.get('date')}
    _cmis = {(d, c) for c, fld, tol in vc.COLS for d in set(_cref) & set(_ceng)
             if vc.num(_cref[d].get(c)) is not None
             and (_ceng[d].get(fld) is None or abs(_ceng[d][fld] - vc.num(_cref[d].get(c))) > tol)}
    check("calculator-page parity: engine matches 'exel caalculator' exactly", not _cmis,
          f"({sorted(_cmis)[:6]})")

print()
if FAILS:
    print(f"{len(FAILS)} FAILED: {FAILS}")
    sys.exit(1)
print("ALL TESTS PASSED")
