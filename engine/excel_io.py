"""
Excel Reader & Export — v2
==========================
Read bond parameters from any Excel layout (all sheets are scanned).
Write professional formatted bond reports that can be re-imported.
"""

import io
from datetime import date, datetime
from typing import Dict, List, Tuple, Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine.calculator import (
    parse_date, safe_float, derive_accrued_interest, fmt_date,
    calc_discount_premium, get_summary
)

# ─── LABEL MAP ────────────────────────────────────────────────────────────────

LABEL_MAP = {
    'par value': 'par_value', 'face value': 'par_value', 'par val': 'par_value',
    'nominal': 'par_value',
    'clean trade price': 'clean_price', 'clean price': 'clean_price',
    'clean tradeprice': 'clean_price', 'purchase price': 'clean_price',
    'coupon rate': 'coupon_rate', 'coupon %': 'coupon_rate',
    'coupon amount': 'coupon_amount',
    'accrued interest': 'accrued_interest', 'accrued': 'accrued_interest',
    'interest frequency': 'interest_frequency', 'coupon frequency': 'interest_frequency',
    'frequency': 'interest_frequency',
    'settle date': 'settle_date', 'settlement date': 'settle_date',
    'last interest date': 'last_interest_date', 'last coupon date': 'last_interest_date',
    'last int': 'last_interest_date',
    'next interest date': 'next_interest_date', 'next coupon date': 'next_interest_date',
    'next int': 'next_interest_date',
    'maturity date': 'maturity_date', 'maturity': 'maturity_date',
    'discount': 'discount', 'premium': 'premium',
    'default probability': 'default_probability', 'drsk': 'default_probability',
    'default prob': 'default_probability', 'probability of default': 'default_probability',
    'lgd': 'lgd', 'loss given default': 'lgd',
    'recovery rate': 'recovery_rate', 'cds recovery': 'recovery_rate',
    'isin': 'isin', 'cusip': 'cusip', 'issuer': 'issuer',
}

# Order matters: 'semi' must be checked before 'annual' ("semi-annual")
FREQ_WORDS = [('semi', 2), ('half', 2), ('quarter', 4), ('month', 12),
              ('annual', 1), ('year', 1)]


# ─── READER ───────────────────────────────────────────────────────────────────

def _norm_text(val) -> str:
    """Lower-case and collapse all whitespace (labels may wrap across lines)."""
    return ' '.join(str(val).split()).lower()


def _scan_blocks(ws) -> List[Dict]:
    """
    Scan one sheet for parameter blocks. A new block starts whenever a label
    already seen in the current block appears again (e.g. a second 'Par value')
    — the layout used to record additional buys of the same bond.
    """
    blocks: List[Dict] = []
    cur: Dict = {}
    seen = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cs = _norm_text(cell.value)
            for label, key in LABEL_MAP.items():
                if label in cs:
                    if key in seen:
                        blocks.append(cur)
                        cur, seen = {}, set()
                    seen.add(key)
                    val = ws.cell(row=cell.row, column=cell.column + 1).value
                    if val is None:
                        val = ws.cell(row=cell.row, column=cell.column + 2).value
                    if val is not None and key not in cur:
                        cur[key] = val
                    break
    if cur:
        blocks.append(cur)
    return blocks


def _block_is_buy(block: Dict) -> bool:
    """A block describes a buy if it has a parseable par, price and settle date."""
    pv = safe_float(block.get('par_value'))
    cp = safe_float(block.get('clean_price'))
    sd = parse_date(block.get('settle_date'))
    return bool(pv and pv > 0 and cp and cp > 0 and sd)


def _parse_tx_type(raw) -> str:
    t = str(raw or 'BUY').upper().strip()
    if 'SELL' in t:
        return 'SELL_FULL' if 'FULL' in t else 'SELL_PARTIAL'
    if t == '1':
        return 'SELL_FULL'
    if t == '2':
        return 'SELL_PARTIAL'
    return 'BUY'


def _scan_transactions(ws, default_price) -> List[Dict]:
    """Find a transaction table on one sheet; returns [] if none."""
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value or 'transaction' not in str(cell.value).lower():
                continue
            hr = cell.row + 1
            col_map = {}
            for hcell in ws.iter_rows(min_row=hr, max_row=hr,
                                      min_col=cell.column, max_col=cell.column + 20):
                for hc in hcell:
                    if not hc.value:
                        continue
                    ht = str(hc.value).strip().lower()
                    if 'date' in ht:              col_map.setdefault('date', hc.column)
                    elif 'nominal' in ht:         col_map['nominal'] = hc.column
                    elif 'price' in ht:           col_map['price'] = hc.column
                    elif 'type' in ht:            col_map['type'] = hc.column
                    elif 'accrued' in ht:         col_map['accrued'] = hc.column
                    elif 'note' in ht:            col_map['note'] = hc.column
            if 'date' not in col_map:
                continue   # not a real transaction table — keep scanning

            txs, blanks, r = [], 0, hr + 1
            base = col_map['date']
            while blanks < 5 and r < hr + 500:
                d = parse_date(ws.cell(row=r, column=col_map['date']).value)
                if not d:
                    blanks += 1
                    r += 1
                    continue
                blanks = 0
                txs.append({
                    'date':             d,
                    'type':             _parse_tx_type(ws.cell(row=r, column=col_map.get('type', base + 4)).value),
                    'nominal':          safe_float(ws.cell(row=r, column=col_map.get('nominal', base + 1)).value) or 0,
                    'clean_price':      safe_float(ws.cell(row=r, column=col_map.get('price', base + 2)).value) or default_price,
                    'accrued_interest': safe_float(ws.cell(row=r, column=col_map.get('accrued', base + 3)).value) or 0,
                    'note':             ws.cell(row=r, column=col_map.get('note', base + 5)).value or '',
                })
                r += 1
            if txs:
                return txs
    return []


def read_from_excel(filepath_or_buffer) -> Tuple[Dict, List[str]]:
    errors = []
    try:
        wb = openpyxl.load_workbook(filepath_or_buffer, data_only=True)
    except Exception as e:
        return {}, [f"Cannot open file: {e}"]

    # Scan sheets in priority order, then any remaining sheets.
    preferred = ['Automated', 'Manual', 'Sheet1', 'Bond', 'Input', 'Parameters']
    ordered = [n for n in preferred if n in wb.sheetnames]
    ordered += [n for n in wb.sheetnames if n not in ordered]

    # Collect parameter blocks across all sheets. The first complete block is
    # the bond itself; every further complete block is an additional BUY of
    # the same bond (par, price, settle date, accrued of its own).
    params: Dict[str, Any] = {}
    extra_buys: List[Dict] = []
    main_found = False
    for name in ordered:
        for block in _scan_blocks(wb[name]):
            if _block_is_buy(block):
                if not main_found:
                    params.update(block)   # bond itself wins over stray fills
                    main_found = True
                else:
                    extra_buys.append({
                        'date':             parse_date(block.get('settle_date')),
                        'type':             'BUY',
                        'nominal':          safe_float(block.get('par_value')),
                        'clean_price':      safe_float(block.get('clean_price')),
                        'accrued_interest': safe_float(block.get('accrued_interest')) or 0,
                        'note':             'Additional buy (parameter block)',
                    })
            else:
                # partial/stray block — fill only keys still missing
                for k, v in block.items():
                    params.setdefault(k, v)

    # Parse dates
    for dk in ['settle_date', 'last_interest_date', 'next_interest_date', 'maturity_date']:
        if dk in params:
            params[dk] = parse_date(params[dk])

    # Frequency may arrive as text ("Semi-annual", "Quarterly", ...)
    if isinstance(params.get('interest_frequency'), str):
        fs = params['interest_frequency'].lower()
        for word, n in FREQ_WORDS:
            if word in fs:
                params['interest_frequency'] = n
                break

    # Safe-float numerics
    for fk in ['par_value', 'clean_price', 'coupon_rate', 'coupon_amount',
               'accrued_interest', 'interest_frequency', 'discount', 'premium',
               'default_probability', 'lgd', 'recovery_rate']:
        if fk in params:
            params[fk] = safe_float(params[fk])
    # (lgd / recovery_rate / default_probability normalization is done by the
    #  engine — fractions and percents are both accepted)

    if not params.get('interest_frequency'):
        params['interest_frequency'] = 2
    params['interest_frequency'] = int(params['interest_frequency'])
    params.setdefault('discount', 0)
    params.setdefault('premium', 0)
    params.setdefault('isin', '')
    params.setdefault('cusip', '')
    params.setdefault('issuer', '')
    for sk in ['isin', 'cusip', 'issuer']:
        params[sk] = str(params[sk] or '').strip()

    # Derive missing values
    if not params.get('coupon_amount') and params.get('par_value') and params.get('coupon_rate'):
        params['coupon_amount'] = params['par_value'] * params['coupon_rate'] / 100.0

    if not params.get('accrued_interest'):
        sd, li, ni = params.get('settle_date'), params.get('last_interest_date'), params.get('next_interest_date')
        pv, cr     = params.get('par_value'), params.get('coupon_rate')
        fr         = params.get('interest_frequency', 2)
        if all([sd, li, ni, pv, cr]):
            params['accrued_interest'] = derive_accrued_interest(pv, cr, fr, sd, li, ni)

    # Scan all sheets for a transaction table; first hit wins.
    transactions: List[Dict] = []
    for name in ordered:
        transactions = _scan_transactions(wb[name], params.get('clean_price', 100))
        if transactions:
            break

    params['transactions'] = extra_buys + transactions
    params['coupon_dates'] = []

    required = ['par_value', 'clean_price', 'coupon_rate', 'settle_date',
                'last_interest_date', 'next_interest_date', 'maturity_date']
    missing = [k for k in required if not params.get(k)]
    if missing:
        errors.append(f"Missing required fields: {', '.join(missing)}")
    else:
        if params['par_value'] <= 0:
            errors.append("Par value must be positive.")
        if params['clean_price'] <= 0:
            errors.append("Clean price must be positive.")
        if params['maturity_date'] <= params['settle_date']:
            errors.append("Maturity date must be after the settlement date.")
        if params['next_interest_date'] <= params['last_interest_date']:
            errors.append("Next interest date must be after the last interest date.")

    return params, errors


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def _fill(color): return PatternFill('solid', start_color=color, end_color=color)
def _border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

FIN  = '#,##0.00;(#,##0.00);"-"'
DATE = 'DD-MMM-YY'
PCT8 = '0.00000000'
NAVY = '1F4E79'; BLUE = '2E75B6'; ALT = 'EBF3FB'
GOLD = 'FFF2CC'; GREEN_BG = 'E2EFDA'; RED_BG = 'FDECEA'; BUY_BG = 'D6E4F0'


def export_to_excel(params: Dict, rows: List[Dict]) -> bytes:
    # Surface the (possibly engine-derived) credit inputs on the export
    params = dict(params)
    if rows:
        if safe_float(params.get('default_probability')) is None:
            params['default_probability'] = rows[0].get('default_prob')
        if safe_float(params.get('lgd')) is None:
            params['lgd'] = rows[0].get('lgd')
    wb = Workbook()
    _write_report(wb, params, rows)
    _write_summary(wb, params, rows)
    _write_params(wb, params)
    _write_transactions(wb, params.get('transactions', []))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_report(wb, params, rows):
    ws = wb.active
    ws.title = 'Bond Report'

    widths = [14,14,14,8,16,20,14,18,20,16,14,12,12,12,14,10,14,18,14,12,10,14,10,14,12,18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ncols = len(widths)
    last_col = get_column_letter(ncols)

    disc, prem = calc_discount_premium(params['par_value'], params['clean_price'])
    bond_type  = 'Discount' if disc < 0 else ('Premium' if prem > 0 else 'Par')

    # Title row
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value     = 'BOND AMORTIZATION & CASHFLOW SCHEDULE  —  MULTI-TRANSACTION'
    c.font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    c.fill      = _fill(NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'].value = f"ISIN: {params.get('isin','—')}   |   Issuer: {params.get('issuer','—')}   |   Type: {bond_type}   |   Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].font  = Font(name='Arial', size=9, italic=True, color='D9E1F2')
    ws['A2'].fill  = _fill(NAVY)
    ws['A2'].alignment = Alignment(horizontal='left', indent=2, vertical='center')
    ws.row_dimensions[2].height = 14

    # Table title
    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'].value = 'CASHFLOW & AMORTIZATION SCHEDULE'
    ws['A3'].font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    ws['A3'].fill  = _fill(BLUE)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 18

    headers = ['Date / Event','Cashflow','Nominal Δ','Days',
               'Bond Discount','Cum Amort Disc',
               'Bond Premium','Cum Amort Prem',
               'Ending Carrying Value','Accrued Interest',
               'Nominal Balance','Price',
               'MTM','OCI G/L','NAV','Check',
               'P&L (Sell)','Realized Interest Income','Total P&L',
               'WAC','Check 2',
               'Default probability (DRSK)','LGD (1-CDS recovery)','Expected Loss','Change',
               'Transaction Detail']
    hr = 4
    ws.row_dimensions[hr].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill      = _fill(BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()

    for i, row in enumerate(rows):
        r = hr + 1 + i
        ws.row_dimensions[r].height = 14

        if row.get('is_maturity'):   bg = GREEN_BG
        elif row.get('is_sell'):     bg = RED_BG
        elif row.get('is_buy') and not row.get('is_header'): bg = BUY_BG
        elif row.get('is_header'):   bg = 'D6E4F0'
        elif row.get('is_coupon'):   bg = GOLD
        elif i % 2 == 1:             bg = ALT
        else:                        bg = 'FFFFFF'

        label = row.get('label') or (fmt_date(row['date']) if row.get('date') else '—')
        # Transactions are recognised on their date, so the carrying value column
        # always shows the position AFTER the trade (remaining holding).
        cv_display = row.get('carrying_value')
        vals = [
            label, row.get('cashflow'), row.get('nominal_change'),
            row.get('num_days'),
            row.get('bond_discount'), row.get('cum_amort_disc'),
            row.get('bond_premium'),  row.get('cum_amort_prem'),
            cv_display, row.get('accrued_int'),
            row.get('nominal_balance'), row.get('price'),
            row.get('mtm'), row.get('oci_gl'), row.get('nav'),
            row.get('check'),
            row.get('realized_pl'),
            row.get('realized_interest_income'), row.get('total_pl'),
            row.get('wac'), row.get('check2'),
            row.get('default_prob'), row.get('lgd'),
            row.get('expected_loss'), row.get('el_change'),
            row.get('tx_detail', ''),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font   = Font(name='Arial', size=9)
            c.fill   = _fill(bg)
            c.border = _border()
            c.alignment = Alignment(horizontal='left' if ci in (1, ncols) else 'right')
            if ci in (2,3,5,6,7,8,9,10,11,13,14,15,16,17,18,19,21,24,25): c.number_format = FIN
            elif ci in (12,20): c.number_format = PCT8
            elif ci == 22: c.number_format = '0.00000000'
            elif ci == 23: c.number_format = '0.00'
            if ci == 5 and isinstance(val,(int,float)) and val < 0:
                c.font = Font(name='Arial', size=9, color='C00000', bold=True)
            if ci == 7 and isinstance(val,(int,float)) and val > 0:
                c.font = Font(name='Arial', size=9, color='375623', bold=True)
            if ci in (17,19) and isinstance(val,(int,float)) and val != 0:
                c.font = Font(name='Arial', size=9, bold=True,
                              color='375623' if val >= 0 else 'C00000')
            # Sell row: red for carrying value (-cv_sold), cashflow, NAV
            if row.get('is_sell') and ci in (2, 9, 14, 15):
                c.font = Font(name='Arial', size=9, color='C00000', bold=True)
            if row.get('is_buy') and ci == 2:
                c.font = Font(name='Arial', size=9, color='375623', bold=True)

    ws.freeze_panes = ws.cell(row=5, column=2)

    # ── Nominal Tracker mini-table (right of main headers, like H2:I7 in reference) ──
    tc = ncols + 2   # first tracker column (gap of 1)
    tracker_events = [
        (r.get('label') or fmt_date(r.get('date')) or '—',
         r.get('nominal_balance'),
         r.get('wac'))
        for r in rows if r.get('is_buy') or r.get('is_sell')
    ]
    if tracker_events:
        # Header
        for ci, h in enumerate(['Transaction', 'Nominal Balance', 'Running WAC'], tc):
            c = ws.cell(row=hr, column=ci, value=h)
            c.font      = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            c.fill      = _fill(NAVY)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = _border()
        ws.column_dimensions[get_column_letter(tc)].width   = 38
        ws.column_dimensions[get_column_letter(tc+1)].width = 18
        ws.column_dimensions[get_column_letter(tc+2)].width = 14
        # Data rows
        for ri, (lbl, nom, wac_v) in enumerate(tracker_events, start=hr + 1):
            bg_t = BUY_BG if 'Buy' in (lbl or '') else RED_BG
            c0 = ws.cell(row=ri, column=tc, value=lbl)
            c0.font = Font(name='Arial', size=9); c0.fill = _fill(bg_t); c0.border = _border()
            c0.alignment = Alignment(horizontal='left')
            c1 = ws.cell(row=ri, column=tc+1, value=nom)
            c1.font = Font(name='Arial', size=9); c1.fill = _fill(bg_t); c1.border = _border()
            c1.alignment = Alignment(horizontal='right'); c1.number_format = FIN
            c2 = ws.cell(row=ri, column=tc+2, value=wac_v)
            c2.font = Font(name='Arial', size=9); c2.fill = _fill(bg_t); c2.border = _border()
            c2.alignment = Alignment(horizontal='right'); c2.number_format = PCT8


def _write_summary(wb, params, rows):
    s = get_summary(params, rows)
    ws = wb.create_sheet('Summary')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22

    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = 'POSITION SUMMARY'
    c.font  = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    c.fill  = _fill(NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    items = [
        ('ISIN',                  params.get('isin', '—') or '—'),
        ('Issuer',                params.get('issuer', '—') or '—'),
        ('Bond Type',             s['bond_type']),
        ('Status',                s['status']),
        ('Initial Discount',      s['discount']),
        ('Initial Premium',       s['premium']),
        ('Initial Purchase Cost', s['purchase_cost']),
        ('YTM at Purchase (%)',   s['ytm']),
        ('Current Yield (%)',     s['current_yield']),
        ('Annual Coupon Income',  s['annual_income']),
        ('Coupon Payments',       s['coupon_payments']),
        ('Years to Maturity',     round(s['years_to_maturity'], 2)),
        ('Total Buys',            s['total_buys']),
        ('Total Sells',           s['total_sells']),
        ('Total Invested',        s['total_invested']),
        ('Total Sell Proceeds',   s['total_sell_proceeds']),
        ('Total Coupon Income',   s['total_coupon_income']),
        ('Redemption at Maturity',s['redemption']),
        ('Net Cashflow',          s['net_cashflow']),
        ('P&L from Sells',        s['total_realized_pl']),
        ('Realized Interest Income', s['realized_interest_income']),
        ('Total Realized P&L',    s['total_pl']),
        ('Default Probability (DRSK)', s['default_probability']),
        ('LGD (1-CDS recovery)',  s['lgd']),
        ('Expected Loss (PD×LGD×Cost)', s['expected_loss']),
        ('Current Nominal Held',  s['current_nominal']),
    ]
    for i, (lbl, val) in enumerate(items, 2):
        lc = ws.cell(row=i, column=1, value=lbl)
        lc.font   = Font(name='Arial', bold=True, size=10, color=NAVY)
        lc.border = _border()
        lc.fill   = _fill(ALT) if i % 2 == 0 else _fill('FFFFFF')
        vc = ws.cell(row=i, column=2, value=val if val is not None else '—')
        vc.font   = Font(name='Arial', size=10)
        vc.border = _border()
        vc.alignment = Alignment(horizontal='right')
        if isinstance(val, float):
            if abs(val) >= 100:  vc.number_format = FIN
            elif abs(val) < 0.01: vc.number_format = '0.00000000'
            else:                vc.number_format = '0.0000'


def _write_params(wb, params):
    ws = wb.create_sheet('Parameters')
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 22
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = 'BOND INPUT PARAMETERS'
    c.font  = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    c.fill  = _fill(NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    disc, prem = calc_discount_premium(params['par_value'], params['clean_price'])
    # Labels match LABEL_MAP so an exported file can be re-imported.
    items = [
        ('ISIN', params.get('isin','—')), ('CUSIP', params.get('cusip','—')),
        ('Issuer', params.get('issuer','—')),
        ('Par Value', params.get('par_value')), ('Clean Price', params.get('clean_price')),
        ('Coupon Rate (%)', params.get('coupon_rate')),
        ('Frequency', params.get('interest_frequency', 2)),
        ('Accrued Interest', params.get('accrued_interest', 0)),
        ('Settle Date', params.get('settle_date')),
        ('Last Interest Date', params.get('last_interest_date')),
        ('Next Interest Date', params.get('next_interest_date')),
        ('Maturity Date', params.get('maturity_date')),
        ('Discount', disc), ('Premium', prem),
        ('Default Probability (DRSK)', params.get('default_probability', 0) or 0),
        ('LGD (1-CDS recovery)', params.get('lgd') if params.get('lgd') is not None else 0.6),
    ]
    for i, (lbl, val) in enumerate(items, 2):
        lc = ws.cell(row=i, column=1, value=lbl)
        lc.font   = Font(name='Arial', bold=True, size=10, color=NAVY)
        lc.border = _border()
        lc.fill   = _fill(ALT) if i % 2 == 0 else _fill('FFFFFF')
        vc = ws.cell(row=i, column=2, value=val)
        vc.font   = Font(name='Arial', size=10, color='0000FF')
        vc.border = _border()
        vc.fill   = _fill(GOLD)
        vc.alignment = Alignment(horizontal='right')
        if isinstance(val, date): vc.number_format = DATE
        elif isinstance(val, float) and abs(val) > 100: vc.number_format = FIN
        elif isinstance(val, float): vc.number_format = PCT8


def _write_transactions(wb, transactions):
    if not transactions:
        return
    ws = wb.create_sheet('Transactions')
    for w, col in zip([14,14,16,16,16,30], 'ABCDEF'):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = 'TRANSACTION LOG'
    c.font  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill  = _fill(NAVY)
    c.alignment = Alignment(horizontal='center')

    hdrs = ['Date','Type','Nominal','Clean Price (%)','Accrued Interest','Note']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = _fill(BLUE); c.border = _border()
        c.alignment = Alignment(horizontal='center')

    for i, tx in enumerate(transactions, 3):
        bg = RED_BG if 'SELL' in str(tx.get('type','')).upper() else BUY_BG
        d = parse_date(tx.get('date'))
        vals = [
            d, str(tx.get('type','BUY')),
            tx.get('nominal',0), tx.get('clean_price',0),
            tx.get('accrued_interest',0), tx.get('note',''),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.font = Font(name='Arial', size=9)
            c.fill = _fill(bg); c.border = _border()
            if ci == 1 and isinstance(val, date): c.number_format = DATE
            if ci in (3,4,5): c.number_format = FIN
