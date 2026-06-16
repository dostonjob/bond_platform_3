"""
Calculator-page verification harness
=====================================
Diffs the engine against the clean, automated 'exel caalculator' sheet in
`For Python_final.xlsx` — a premium bond (150M @ 108.1, 6.25% semi, no
transactions) with input blocks (params A1:B14, coupon dates in col F,
transaction table in J:V) and an auto-computed schedule from row 17.

This is the canonical reference: the engine must reproduce every valuation
column exactly. Run:  set PYTHONIOENCODING=utf-8 && python verify_calc.py
"""
import sys, os
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from engine.calculator import build_cashflow_table

REF_FILE = 'For Python_final.xlsx'
SHEET    = 'exel caalculator'

# calculator-page column letter -> (engine row field, tolerance)
COLS = [
    ('D', 'num_days',        1.0),
    ('E', 'bond_discount',   1.0),
    ('F', 'cum_amort_disc',  1.0),
    ('G', 'bond_premium',    1.0),
    ('H', 'cum_amort_prem',  1.0),   # NEGATIVE for premium (amortizes carrying down)
    ('J', 'carrying_value',  1.0),
    ('K', 'accrued_int',     1.0),
    ('L', 'nominal_balance', 1.0),
    ('M', 'price',           1e-4),
    ('N', 'mtm',             1.0),
    ('O', 'oci_gl',          1.0),
    ('P', 'nav',             1.0),
    ('Q', 'check',           0.01),
]
COL_IDX = {c: i for i, c in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ', 1)}


def read_reference():
    """Return {date: {col: value}} for the schedule rows + market prices from col M."""
    ws = openpyxl.load_workbook(REF_FILE, data_only=True)[SHEET]
    ref, prices = {}, {}
    for r in range(19, 224):
        a = ws.cell(row=r, column=1).value
        if not isinstance(a, datetime):
            continue
        d = a.date()
        ref[d] = {c: ws.cell(row=r, column=COL_IDX[c]).value for c, _f, _t in COLS}
        m = ws.cell(row=r, column=COL_IDX['M']).value
        if m is not None:
            prices[d] = float(m)
    return ref, prices


def build_params(prices):
    return dict(
        isin='aaa1', par_value=150_000_000.0, clean_price=108.1015625, coupon_rate=6.25,
        interest_frequency=2, settle_date=date(2026, 5, 14), last_interest_date=date(2025, 11, 15),
        next_interest_date=date(2026, 5, 15), maturity_date=date(2030, 5, 15),
        market_prices=prices)


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def main():
    ref, prices = read_reference()
    eng = {r['date']: r for r in build_cashflow_table(build_params(prices)) if r.get('date')}
    common = sorted(set(ref) & set(eng))
    print(f'calculator-page rows: {len(ref)} | engine: {len(eng)} | common: {len(common)}')
    print()
    total = 0
    for c, fld, tol in COLS:
        m = t = 0
        bad = []
        for d in common:
            rv = num(ref[d].get(c))
            if rv is None:
                continue
            ev = num(eng[d].get(fld))
            t += 1
            if ev is not None and abs(ev - rv) <= tol:
                m += 1
            elif len(bad) < 2:
                bad.append(f'{d}: {ev} vs {rv:,.2f}')
        total += (t - m)
        print(f'{"OK " if m == t else "!! "}{c:2} {fld:16} {m:>4}/{t:<4}   {" | ".join(bad)}')
    print()
    print('ALL VALUATION COLUMNS MATCH' if total == 0 else f'{total} mismatched cells')


if __name__ == '__main__':
    main()
